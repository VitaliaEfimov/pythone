from openpyxl import Workbook
from openpyxl.utils.exceptions import IllegalCharacterError

def safe_formula_write(ws, cell, formula):
    """Безопасная запись формулы"""
    try:
        ws[cell] = formula
        return True
    except IllegalCharacterError:
        print(f"Ошибка в формуле: {formula}")
        return False

wb = Workbook()
ws = wb.active

ws['A1'] = 5
ws['A2'] = 10

# Безопасная запись формулы
safe_formula_write(ws, 'A3', '=СУММ(A1:A2)')

wb.save('formula.xlsx')